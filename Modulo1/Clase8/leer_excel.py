import openpyxl

path = "Estudiantes/archivo_estudiantes.xlsx"  # Ruta del archivo Excel
wb = openpyxl.load_workbook(path)  # Carga el libro de trabajo existente
ws = wb.active  # Obtiene la hoja activa del libro de trabajo

print(ws.title)  # Imprime el título de la hoja activa

rows = ws.iter_rows(min_row=2, values_only=True)# Itera sobre las filas de la hoja, comenzando desde la segunda fila

for row in rows:
    print(row)  # Imprime cada fila como una tupla
    nombre, edad, carrera = row  # Desempaqueta los valores de la fila
    print(f"Nombre: {nombre}, Edad: {edad}, Carrera: {carrera}")  # Imprime los valores de cada estudiante
    
# # Insertar una nueva fila al final de la hoja
# ws.append(["Pedro", 24, "Biología"])  # Agrega un nuevo estudiante

# Guardar los cambios en el archivo Excel
wb.save(path)  # Guarda el libro de trabajo en el mismo archivo
print("Archivo actualizado correctamente.")  # Mensaje de confirmación

# Crear una hoja nueva y agregar datos
ws_nueva = wb.create_sheet(title="Profesores")  # Crea una nueva hoja con el título "Profesores"

ws_nueva.append(["Nombre", "Edad", "Materia"])  # Agrega los encabezados de las columnas
ws_nueva.append(["Carlos", 35, "Matemáticas"])  # Agrega un nuevo profesor
ws_nueva.append(["Laura", 30, "Historia"])  #

wb.save(path)  # Guarda los cambios en el archivo Excel

new_ws = wb["Profesores"]  # Obtiene la nueva hoja creada
print(new_ws.title)  # Imprime el título de la nueva hoja
print(new_ws["A1"].value)  # Imprime el valor de la celda A1 de la nueva hoja