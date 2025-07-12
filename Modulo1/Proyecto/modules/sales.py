from openpyxl import Workbook, load_workbook
import os

def export(path, current_time, ventas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Write headers
    ws.append(['Producto', 'Cantidad', 'Cliente'])
    
    # Write data
    for venta in ventas:
        ws.append([venta['producto'], venta['cantidad'], venta['cliente']])
    
    file_name = f'ventas_{current_time.replace(" ", "_").replace(":", "-")}.xlsx'
    file_path = os.path.join(path, file_name)
    wb.save(file_path)
    
def import_sales_mod(path, file):
    file_path = os.path.join(path, file.filename)
    file.save(file_path)
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    ventas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ventas.append({
            'producto': row[0],
            'cantidad': row[1],
            'cliente': row[2]
        })
    
    return ventas